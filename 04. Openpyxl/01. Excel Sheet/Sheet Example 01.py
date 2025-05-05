import openpyxl

from openpyxl import load_workbook

# You must change the physical path before running this script.
currPath = "C:/Users/natio/OneDrive - 성균관대학교/99. Personal Blog/05. Python/05. OPENPYXL/01. Excel Sheet/"

# # Load the excel file
wb = openpyxl.load_workbook(filename=currPath+"Sample Sheet File.xlsx")
res = len(wb.sheetnames)

# Number of sheets
print('Number of sheets : ', res, 'sheets')

ws = wb.active

# print the sheet name
print('Active sheet name : ', ws.title)




# #셀 읽는 2가지 방법
# print(sheet['A1'].value)
# print(sheet.cell(row=1,column=1).value)



# # #Create a workbook
# # wb = openpyxl.Workbook()
# # ws = wb.active

# # ws1 = wb.create_chartsheet("Mysheet")       # insert at the end (default)
# # ws2 = wb.create_chartsheet("Mysheet", 0)    # insert at first position
# # ws3 = wb.create_chartsheet("Mysheet", -1)   # insert at the penultimate position

