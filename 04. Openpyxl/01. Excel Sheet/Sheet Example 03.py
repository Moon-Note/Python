import openpyxl

from openpyxl import workbook

# You must change the physical path before running this script.
currPath = "C:/Users/natio/OneDrive - 성균관대학교/99. Personal Blog/05. Python/05. OPENPYXL_XLRD_XLWT/01. Excel Sheet/"

# Load the excel file
wb = openpyxl.load_workbook(filename=currPath+"Sample Sheet File.xlsx")

# Create New Sheet
#wb.create_sheet(title='New Sheet', index=0)

# Copy Sheet
copy_Sheet = wb.copy_worksheet(wb["Sample Sheet"])

# Copied sheet Name is added 'Copy' at the end of the name
print(wb.sheetnames)

# Change the copied sheet name
copy_Sheet.title = "Duplicate Sheet"

# Confirm the Sheet names
print(wb.sheetnames)