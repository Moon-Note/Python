from openpyxl import Workbook
from openpyxl.styles import Protection

# 워크북 및 시트 생성
wb = Workbook()
ws = wb.active

# 셀 잠금 설정 (Protection 객체를 사용하여 설정)
ws["A1"].protection = Protection(locked=True)   # A1 셀 잠금
ws["B1"].protection = Protection(locked=False)  # B1 셀은 잠금 해제

# 시트 보호 설정 (비밀번호 필요)
ws.protection.set_password("1234")
ws.protection.enable()  # 시트 보호 활성화

# 파일 저장
wb.save("cell_lock_example.xlsx")
