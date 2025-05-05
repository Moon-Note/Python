from openpyxl import Workbook
from openpyxl.styles import Protection

# 워크북 및 시트 생성
wb = Workbook()
ws = wb.active

# 시트 전체의 모든 셀을 잠금 해제 (기본적으로 엑셀 시트는 잠금 상태로 생성됨)
for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
    for cell in row:
        cell.protection = Protection(locked=False)

# 특정 열에 대해 보호 속성 설정
for row in range(1, 11):
    ws[f"A{row}"].protection = Protection(locked=False)   # A 열은 잠금 해제
    ws[f"B{row}"].protection = Protection(locked=True)    # B 열은 잠금

# 시트 보호 설정 (비밀번호 필요)
ws.protection.set_password("1234")
ws.protection.enable()  # 시트 보호 활성화

# 파일 저장
wb.save("advanced_cell_lock_example.xlsx")
