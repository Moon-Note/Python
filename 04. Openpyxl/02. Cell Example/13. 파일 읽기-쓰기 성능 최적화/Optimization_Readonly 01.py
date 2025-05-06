from openpyxl import load_workbook
import os

# You must change the physical path before running this script.
currPath = os.getcwd()+"/04. Openpyxl/02. Cell Example/13. 파일 읽기-쓰기 성능 최적화/"
# # # Check the current path
print("Current Path:", currPath)

wb = load_workbook(filename=currPath+"large_file.xlsx", read_only=True)
ws = wb.active

for row in ws.iter_rows(min_row=1, max_col=3, max_row=10):
   print([cell.value for cell in row])

