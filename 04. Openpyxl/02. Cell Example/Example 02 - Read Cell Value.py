import openpyxl
import os

from openpyxl import load_workbook

# You must change the physical path before running this script.
currPath = os.getcwd()+"/04. Openpyxl/02. Cell Example/"
# # Check the current path
print("Current Path:", currPath)


# # Load the excel file
wb = openpyxl.load_workbook(filename=currPath+"Cell Example File.xlsx")
ws = wb.active

# initialize row index
x = 0
# Result data
data = []

for i in range(5):
    x += 1
    y = 0
    R_data = []

    for j in range(5):
        y += 1

        R_data.append(ws.cell(x,y).value)
    data.append(R_data)

# Excel 2D Data Array
print(data)



