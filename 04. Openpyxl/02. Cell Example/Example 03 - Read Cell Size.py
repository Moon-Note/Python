import openpyxl

from openpyxl import load_workbook

# You must change the physical path before running this script.
currPath = "C:/Users/natio/OneDrive - 성균관대학교/99. Personal Blog/05. Python/05. OPENPYXL/02. Cell Example/"

# # Load the excel file
wb = openpyxl.load_workbook(filename=currPath+"Cell Example File.xlsx")
ws = wb.active

print('data size[row, column]')
#Method 01 : max_row & max_column commands
print('row : ', ws.max_row, ', column : ', ws.max_column)

#Method 02 : len(ws['*specific cell name']) 
print('A column size : ', len(ws['A']))