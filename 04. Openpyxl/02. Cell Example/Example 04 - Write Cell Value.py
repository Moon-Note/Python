import openpyxl

from openpyxl import load_workbook

# You must change the physical path before running this script.
currPath = "C:/Users/natio/OneDrive - 성균관대학교/99. Personal Blog/05. Python/05. OPENPYXL/02. Cell Example/"

# # Load the excel file
wb = openpyxl.load_workbook(filename=currPath+"Cell Write Example File.xlsx")
ws = wb.active

# Read Method 01 : ['Cell'].value
print('Initial A1 Value : ', ws['A1'].value)
# Read Method 02 : cell(row=x, column=y).value
print('Initial B1 Value : ', ws.cell(row=1,column=2).value)

# Write Method 01 : ['Cell'].value
ws['A1'].value = 111
# Write Method 02 : cell(row=x, column=y).value
ws.cell(row=1,column=2).value = 222


# Read Method 01 : ['Cell'].value
print('Modified A1 Value : ', ws['A1'].value)
# Read Method 02 : cell(row=x, column=y).value
print('Modified B1 Value : ', ws.cell(row=1,column=2).value)

