import openpyxl

from openpyxl import load_workbook

# You must change the physical path before running this script.
currPath = "C:/Users/natio/OneDrive - 성균관대학교/99. Personal Blog/05. Python/05. OPENPYXL/02. Cell Example/"

# # Load the excel file
wb = openpyxl.load_workbook(filename=currPath+"Cell Example File.xlsx")
ws = wb.active

# initialize row index
x = 0
# Result data
data = []

for i in range(5):
    x += 1
    y = 0
    W_data = []

    for j in range(5):
        y += 1
        
        ws.cell(x,y).value = 5*(x-1)+y
        CellData = ws.cell(x,y).value
        W_data.append(CellData)
    data.append(W_data)

# Excel 2D Data Array
print(data)



