from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame

# 새 워크북 생성
workbook = Workbook()

# 현재 활성화된 시트 선택
sheet = workbook.active

# 중복 데이터를 포함하는 DataFrame 생성
data = {'Name': ['Alice', 'Bob', 'Alice', 'Charlie', 'Bob', 'Alice'],
        'Age': [25, 30, 25, 35, 30, 25],
        'City': ['New York', 'Paris', 'London', 'Tokyo', 'Paris', 'New York']}
df = DataFrame(data)

# DataFrame의 데이터를 시트에 작성
for row in dataframe_to_rows(df, index=False, header=True):
    sheet.append(row)

# 중복 데이터를 제거하는 과정
rows = list(sheet.iter_rows(values_only=True))
seen = set()
rows_without_duplicates = []
for row in rows:
    if row not in seen:
        seen.add(row)
        rows_without_duplicates.append(row)

# 시트 초기화
sheet.delete_rows(1, sheet.max_row)

# 중복이 제거된 데이터를 시트에 작성
for row in rows_without_duplicates:
    sheet.append(row)

# 워크북 저장
workbook.save('[Exam 01]remove_duplicates_by_row.xlsx')
