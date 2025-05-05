from openpyxl import load_workbook

def search_cell_value(workbook_path, sheet_name, target_value):
    # 엑셀 파일 불러오기
    wb = load_workbook(filename=workbook_path)
    
    # 시트 선택
    sheet = wb[sheet_name]

    # 결과를 저장할 딕셔너리 초기화
    result = {'value': target_value, 'count': 0, 'positions': []}

    # 시트 내 모든 셀 순회
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            # 셀의 값이 타겟 값과 일치하는지 확인
            if cell.value == target_value:
                # 타겟 값이 일치할 경우 개수 증가 및 위치 정보 저장
                result['count'] += 1
                result['positions'].append((row, column))

    return result

def replace_cell_value(workbook_path, sheet_name, target_value, replacement, new_workbook_path):
    # 엑셀 파일 불러오기
    wb = load_workbook(filename=workbook_path)
    
    # 시트 선택
    sheet = wb[sheet_name]

    # 시트 내 모든 셀 순회
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            # 셀의 값이 타겟 값과 일치하는지 확인
            if cell.value == target_value:
                # 타겟 값이 일치할 경우 대체값으로 변경
                cell.value = replacement

    # 변경된 내용을 저장
    wb.save(new_workbook_path)


# You must change the physical path before running this script.
currPath = "C:/Users/natio/OneDrive/0. Personal Blog/05. Python/05. OPENPYXL/02. Cell Example/"
# 예시: 특정 엑셀 파일에서 '!' 값을 검색하고 '@'로 대체
workbook_path = currPath+"Find Cell Value Example File.xlsx"
print('workbook_path :', workbook_path)
sheet_name = 'Sample Sheet'
target_value = '!'
replacement = '@'
#새로운 파일 이름 지정
new_workbook_path = currPath+"Replaced Cell Value Example File.xlsx"

# 검색한 값 대체
replace_cell_value(workbook_path, sheet_name, target_value, replacement, new_workbook_path)
# 결과 확인
result = search_cell_value(workbook_path, sheet_name, target_value)

print(f"검색 문자 : {result['value']}, 대체 문자 : {replacement}")
print(f"총 개수: {result['count']}")
print("위치 정보:")
for position in result['positions']:
    print(f"행: {position[0]}, 열: {position[1]}")

