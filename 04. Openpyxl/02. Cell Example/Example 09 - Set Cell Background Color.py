from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 새 워크북 생성
wb = Workbook()

# 활성 시트 선택
ws = wb.active

# 데이터 입력
ws['A1'] = 'Hello'
ws['B2'] = 'World'

# 배경색 지정 (빨강)
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# 배경색 지정 (파랑)
blue_fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')

# 셀에 배경색 적용
ws['A1'].fill = red_fill
ws['B2'].fill = blue_fill

# 파일 저장
wb.save("colored_cells.xlsx")
