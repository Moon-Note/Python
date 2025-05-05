from openpyxl import Workbook
from openpyxl.styles import Font

# 새 워크북 생성
wb = Workbook()

# 활성 시트 선택
ws = wb.active

# 데이터 입력
ws['A1'] = 'MoonNote'
ws['B2'] = 'Tistory'
ws['C3'] = 'Blog'

# 글꼴 스타일 지정
bold_font = Font(name='Arial', size=14, bold=True)
italic_font = Font(name='Times New Roman', size=12, italic=True)
underline_font = Font(name='Calibri', size=16, underline='single')
color_font = Font(name='Verdana', size=10, color='FF0000')  # 빨간색

# 셀에 글꼴 스타일 적용
ws['A1'].font = bold_font
ws['B2'].font = italic_font
ws['C3'].font = underline_font
ws['C3'].font = color_font

# 파일 저장
wb.save("styled_fonts.xlsx")
