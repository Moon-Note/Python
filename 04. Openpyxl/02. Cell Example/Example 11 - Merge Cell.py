from openpyxl import Workbook

# 새 워크북 생성
wb = Workbook()

# 활성 시트 선택
ws = wb.active

# 데이터 입력
ws['A1'] = 'Merged Cell'

# 셀 병합
ws.merge_cells('A1:C1')  # A1에서 C1까지 병합
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)  # (2,1)에서 (2,3)까지 병합

# 병합된 셀에 데이터 입력 (병합된 셀의 좌상단 셀에만 값을 넣어야 함)
ws['A2'] = 'Merged Cell 2'

# 파일 저장
wb.save("merged_cells.xlsx")
