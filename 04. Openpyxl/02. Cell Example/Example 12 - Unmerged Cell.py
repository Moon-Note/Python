from openpyxl import load_workbook

# 기존 워크북 불러오기
wb = load_workbook("merged_cells.xlsx")

# 활성 시트 선택
ws = wb.active

# 셀 분할
ws.unmerge_cells('A1:C1')  # A1에서 C1까지 병합 해제
ws.unmerge_cells(start_row=2, start_column=1, end_row=2, end_column=3)  # (2,1)에서 (2,3)까지 병합 해제

# 파일 저장
wb.save("unmerged_cells.xlsx")