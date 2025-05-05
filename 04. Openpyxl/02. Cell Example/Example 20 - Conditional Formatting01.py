from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

# 새 워크북과 워크시트 생성
wb = Workbook()
ws = wb.active

# 예제 데이터 추가
data = [
    ["학생", "점수"],
    ["Alice", 75],
    ["Bob", 85],
    ["Charlie", 95],
    ["David", 70]
]

for row in data:
    ws.append(row)

# 조건부 서식 규칙 설정: 점수가 80 이상일 때 셀 배경을 초록색으로 변경
fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
rule = CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=fill)
ws.conditional_formatting.add("B2:B5", rule)

# 워크북 저장
wb.save("conditional_formatting_example01.xlsx")
